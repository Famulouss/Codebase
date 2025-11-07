import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import scipy.signal as signal
import tkinter as tk
from scipy.signal import butter, filtfilt
from tkinter import Tk, ttk, filedialog
from tkinter.filedialog import askopenfilename
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
import seaborn as sns

# -------------------
# Globale Variablen
# -------------------
delta_time = 0.5
used_interval = 5.0
unkomfortabel_detektiert = {'M1': [], 'M2': [], 'M3': []}
# =====================================================
# UNKOMFORTABLE-ZEITEN EINTRAGEN (manuell anpassbar)
# =====================================================
# Liste mit Start- und Endzeiten (Sekunden)
unangenehme_zeiten = [14, 99, 167, 194, 278, 308, 338, 358, 411, 458, 468, 519, 524, 529, 585, 663, 740, 755]
unangenehme_intervalle = []
for i, time in enumerate(unangenehme_zeiten):
    unangenehme_intervalle.append((time-delta_time, time+delta_time))


# m1: Methode 1 -> Einfache Threshold-Detektion (Threshold aus Paper)
# m2: Methode 2 -> Thresholds für Integral über unterschiedlich große Intervalle (1s, 3s, 5s) (Thresholds müssen noch bestimmt werden)
# m3: Methode 3 -> Thresholds für RMS nach Norm 2631
norm_thresh = [0.315, 0.5, 0.63, 0.8, 1.25, 1.6, 2, 2.5]
thresholds_acc_x = {'m1': 2.25, 'm2': [2.49, 5.43, 8.85], 'm3':norm_thresh}
thresholds_acc_y = {'m1': 3.44, 'm2': [3.38, 9.61, 13.1], 'm3':norm_thresh}
thresholds_jerk_x = {'m1': 3.0, 'm2': [2.72, 6.37, 7.92]}
thresholds_jerk_y = {'m1': 4.15, 'm2': [3.44, 5.76, 6.52]}
thresholds_acc_yaw = {'m1': 0.45, 'm2': [0.39, 0.8, 0.86], 'm3':norm_thresh}
thresholds_yaw_rate = {'m1': 0.81}
thresholds_dict = {'acc_x': thresholds_acc_x, 'acc_y': thresholds_acc_y, 'acc_yaw': thresholds_acc_yaw, 'jerk_x': thresholds_jerk_x, 'jerk_y': thresholds_jerk_y, 'yaw_rate': thresholds_yaw_rate}

# -------------------
# Hilfsfunktionen
# -------------------
def confirm_selection():    # Für die Abtastratenauswahl
    value = float(combo.get())
    global fs_glob
    fs_glob = 1000/value
    print(f"Gewählte Abtastrate: {value} ms, {fs_glob} Hz")
    root.destroy()

def butter_lowpass(cutoff, fs, order=2):
    """Butterworth Tiefpass-Filter"""
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = butter(order, normal_cutoff, btype='low', analog=False)
    return b, a

def lowpass_filter(data, cutoff, fs, order=2):
    """Tiefpass-Filter für Butterworth-Filter"""
    b, a = butter_lowpass(cutoff, fs, order=order)
    y = filtfilt(b, a, data)
    return y

def integral_abs(times, signal):
    """Integral der absoluten Beschleunigung: ∫ |a(t)| dt (Trapezregel)."""
    return np.trapezoid(np.abs(signal), times)

def highpass_norm(omega1):
    """
    Aufteilung in Numerator und Denominator des Hochpassfilters aus der Norm 2631: 
    Hh(p) = 1/(1+sqrt(2)*omega1/p+(omega1/p)^2)
    """
    num = [1, 0, 0] # Numerator (Zählerkoeffizienten N(s))
    den = [1, np.sqrt(2)*omega1, omega1**2] # Denominator (Nennerkoeffizienten (D(s))
    return num, den

def lowpass_norm(omega2):
    """
    Aufteilung in Numerator und Denominator des Tiefpassfilters aus der Norm 2631:
    Hl(p) = 1/(1+sqrt(2)*p/omega2+(p/omega2)^2)
    """
    num = [1]
    den = [1/(omega2**2), np.sqrt(2)/omega2, 1]
    return num, den

def acc_vel_transition_filter(omega3, omega4, Q4):
    """
    Aufteilung in Numerator und Denominator des "acceleration-velocity transition"-Filters aus der Norm 2631:
    Hl(p) = (1+p/omega3)/(1+p/(Q4*omega4)+(p/omega4)^2)
    """
    num = [1/omega3, 1]
    den = [1/(omega4**2), 1/(Q4*omega4), 1]
    return num, den

def design_wd_filter(fs):
    """
    Entwirft digitale Filterkoeffizienten für Wd-Gewichtung gemäß ISO 2631 (Annäherung).
    fs: Abtastrate in Hz
    returns: b, a (arrays)
    """
    omega1 = 2 * np.pi * 0.4    # f1 = 0,4 Hz
    omega2 = 2 * np.pi * 100.0  # f2 = 100 Hz
    omega3 = 2 * np.pi * 2.0    # f3 = 2 Hz
    omega4 = 2 * np.pi * 2.0    # f4 = 2 Hz
    q4 = 0.63

    # Konstruktion der Teilfilter (analog)
    num_h, den_h = highpass_norm(omega1)    # Hochpassfilter
    num_l, den_l = lowpass_norm(omega2)   # Tiefpassfilter
    num_t, den_t = acc_vel_transition_filter(omega3, omega4, q4)    # Acceleration-velocity transition Filter
    
    # Gesamtnumerator/-denominator durch Polynom-Multiplikation (Faltung der Koeffizienten)
    num_temp = np.convolve(num_h, num_l)
    num_total = np.convolve(num_temp, num_t)
    den_temp = np.convolve(den_h, den_l)
    den_total = np.convolve(den_temp, den_t)

    # Diskretisierung/Umwandlung in digitales Filter (bilineare Transformation)
    b, a = signal.bilinear(num_total, den_total, fs)
    
    return b, a

def design_we_filter(fs):
    """
    Entwirft digitale Filterkoeffizienten für We-Gewichtung gemäß ISO 2631 (Annäherung).
    fs: Abtastrate in Hz
    returns: b, a (arrays)
    """
    omega1 = 2 * np.pi * 0.4    # f1 = 0,4 Hz
    omega2 = 2 * np.pi * 100.0  # f2 = 100 Hz
    omega3 = 2 * np.pi * 1.0    # f3 = 1 Hz
    omega4 = 2 * np.pi * 1.0    # f4 = 1 Hz
    q4 = 0.63

    # Konstruktion der Teilfilter (analog)
    num_h, den_h = highpass_norm(omega1)    # Hochpassfilter
    num_l, den_l = lowpass_norm(omega2)   # Tiefpassfilter
    num_t, den_t = acc_vel_transition_filter(omega3, omega4, q4)    # Acceleration-velocity transition Filter
    
    # Gesamtnumerator/-denominator durch Polynom-Multiplikation (Faltung der Koeffizienten)
    num_temp = np.convolve(num_h, num_l)
    num_total = np.convolve(num_temp, num_t)
    den_temp = np.convolve(den_h, den_l)
    den_total = np.convolve(den_temp, den_t)

    # Diskretisierung/Umwandlung in digitales Filter (bilineare Transformation)
    b, a = signal.bilinear(num_total, den_total, fs)

    return b, a

def sliding_metrics(signal, times, window_s=1.0, overlap=0.5, weighted=False):
    """
    Berechnet die gleitenden RMS-Intervalle übers Signal hinweg. Achtung: times muss in Sekunden sein (nicht ms)!
    Weighted=True: Der frequenzgewichtete Mittelwert (RMS) wird pro Intervall berechnet.
    Weighted=False: Es wird das Integral über die absoluten Beschleunigungswerte pro Intervall berechnet.
    Return: rms_list, rms_times
    """
    times = np.asarray(times)
    signal = np.asarray(signal)
    
    step = window_s * (1.0-overlap)
    if step <= 0:
        raise ValueError("Overlap ist zu groß (<1)")
    # Ermittle die Fenstermittelpunkte
    t_start = times[0]
    t_end = times[-1]
    centers = np.arange(t_start + window_s/2.0, t_end - window_s/2.0 + 1e-9, step)  # Von 1. Argument werden im Abstand vom 3. Argument die Werte ermittelt bis zum ausschließlich 2. Argument

    value_list = []

    # Ermittle den RMS Wert für jedes Intervall
    for tc in centers:
        t0 = tc - window_s/2.0
        t1 = tc + window_s/2.0
        mask = (times >= t0) & (times <= t1)
        seg_t = times[mask]
        seg = signal[mask]

        # RMS Berechnung
        if weighted:
            # Integral für den frequenzgewichteten RMS (∫(aw^2)dt)
            integral = np.trapezoid(seg**2, seg_t)
            # weighted r.m.s (sqrt(1/T*I))
            rms = np.sqrt(integral/window_s)
            value_list.append(rms)
        # Integralberechnung
        else:
            integral = integral_abs(seg_t, seg)
            value_list.append(integral)

    return value_list, centers

def threshold_detection_single(signal, times, threshold):
    """
    Berechnet die Anzahl und die Gesamtdauer der Threshold-Überschreitungen mit einem threshold.
    """
    times = np.asarray(times)
    signal = np.asarray(signal)

    # Maske: True, wenn Signal über Schwellwert
    above = signal > threshold

    exceed_durations = []
    in_exceed = False
    start_time = None

    for i in range(len(signal)):
        if above[i] and not in_exceed:
            # Beginn einer Überschreitung
            in_exceed = True
            start_time = times[i]
        elif not above[i] and in_exceed:
            # Ende einer Überschreitung
            end_time = times[i]
            exceed_durations.append(end_time - start_time)
            in_exceed = False
    # Falls das Signal am Ende immer noch über Threshold ist
    if in_exceed and start_time is not None:
        exceed_durations.append(times[-1] - start_time)

    exceed_count = len(exceed_durations)
    total_duration = np.sum(exceed_durations) if exceed_durations else 0.0

    return exceed_count, exceed_durations, total_duration

def get_max_per_interval(time, signal, interval_size=1.0):
    """
    Berechnet den Maximalwert des Signals in jedem Intervall der gegebenen Größe.
    Gibt ein Dictionary mit Intervallzentren als Keys und Maximalwerten als Values zurück.
    """
    time = np.asarray(time)
    signal = np.asarray(signal)

    max_values = {}
    start_time = np.floor(time[0])
    end_time = np.ceil(time[-1])
    
    intervals = np.arange(start_time, end_time, interval_size)
    
    for start in intervals:
        end = start + interval_size
        mask = (time >= start) & (time < end)
        if np.any(mask):
            max_val = np.max(signal[mask])
            center = start + interval_size / 2
            max_values[center] = max_val
        else:
            # Falls kein Wert im Intervall liegt
            max_values[start + interval_size / 2] = np.nan

    return max_values

def timestamp_is_uncomfortable(time_value):
    """Prüft, ob Zeitwert innerhalb eines unkomfortablen Intervalls liegt."""
    # Hole das Label für den Zeitwert (nächster Zeitindex)
    idx = (df['time_rel'] - time_value).abs().idxmin()
    return df.loc[idx, 'discomfort'] == 1

def interval_is_uncomfortable(time_value, interval_length=1.0):
    """
    Prüft, ob innerhalb eines gegebenen Zeitintervalls ein unkomfortabler Zeitstempel liegt.

    Parameter:
        time_value (float): Zentrum des Intervalls (in Sekunden)
        interval_length (float): Länge des Intervalls (z. B. 1.0 für 1 s)
    
    Rückgabe:
        bool: True, wenn im Intervall ein unkomfortabler Wert liegt, sonst False
    """

    # Intervallgrenzen berechnen
    half = interval_length / 2
    start = time_value - half
    end = time_value + half

    # Daten innerhalb des Intervalls auswählen
    mask = (df['time_rel'] >= start) & (df['time_rel'] <= end)
    interval_data = df.loc[mask, 'discomfort']

    # Prüfen, ob irgendein Wert in diesem Intervall unkomfortabel ist
    return (interval_data == 1).any()

def plot_two_normalized_histograms(data1, data2, signal_name, output_dir, bins=40, label1="Komfortabel", label2="Unkomfortabel",
                                   title="Normalisiertes Histogramm", color1='tab:blue', color2='tab:orange'):
    """
    Plottet zwei Histogramme in normalisierter Form (Summe der Balkenhöhen = 1).

    Parameter:
        data1, data2: array-ähnlich
            Die beiden Datensätze (z. B. komfortabel / unkomfortabel)
        signal_name:
            Signalname zur Bestimmung der Achsenbeschriftung
        bins: int oder Sequenz
            Anzahl oder Grenzen der Bins
        label1, label2: str
            Beschriftungen der beiden Histogramme
        title: str
            Titel des Plots
        color1, color2: str
            Farben der beiden Histogramme
    """
    # x-Label bestimmen
    if 'acc_x' in signal_name or 'acc_y' in signal_name:
        xlabel = "Beschleunigunswert [m/s²]"
    elif 'acc_yaw' in signal_name:
        xlabel = "Beschleunigungswert [rad/s²]"
    elif 'jerk_x' in signal_name or 'jerk_y' in signal_name:
        xlabel = "Jerkwert [m/s³]"

    # --- Histogramme berechnen ---
    counts1, bins1 = np.histogram(data1, bins=bins)
    counts2, bins2 = np.histogram(data2, bins=bins)

    # --- Normalisierung (Summe der Balkenhöhen = 1) ---
    counts1 = counts1 / np.sum(counts1) if np.sum(counts1) != 0 else counts1
    counts2 = counts2 / np.sum(counts2) if np.sum(counts2) != 0 else counts2
    
    # --- Plot ---
    fig = plt.figure(figsize=(6, 3))

    plt.hist(data1, bins=bins, alpha=0.6, label=label1, color=color1, weights=np.ones(len(data1))/len(data1))
    plt.hist(data2, bins=bins, alpha=0.6, label=label2, color=color2, weights=np.ones(len(data2))/len(data2))

    plt.xlabel(xlabel)
    plt.ylabel("Relative Häufigkeit (Σ=1)")
    plt.title(title)
    plt.legend()
    plt.tight_layout()
    plt.show()
    fig.savefig(os.path.join(output_dir, f"02_Histogramm_{signal_name}.png"))

def find_threshold_intervals(signal, timestamps, threshold, interval_size=None):
    """
    Findet Zeitintervalle, in denen das Signal den Grenzwert überschreitet.

    Parameter
    ----------
    signal : array-like
        Signalwerte (z. B. numpy array oder Pandas Series)
    timestamps : array-like
        Zeitstempel, gleiche Länge wie signal
    threshold : float
        Grenzwert, ab dem das Signal als 'überschritten' gilt

    Rückgabe
    --------
    intervals : list of tuples
        Liste von (start_time, end_time) Intervallen, 
        die direkt mit plt.axvspan() markiert werden können.
    """
    signal = np.asarray(signal)
    timestamps = np.asarray(timestamps)

    # Bool-Array: True, wenn der absolute Signal-Wert über dem Grenzwert liegt
    above = np.abs(signal) > threshold

    # Übergänge finden
    starts = np.where(np.diff(above.astype(int)) == 1)[0] + 1
    ends = np.where(np.diff(above.astype(int)) == -1)[0] + 1

    # Falls Signal mit Überschreitung startet oder endet, anpassen
    if above[0]:
        starts = np.insert(starts, 0, 0)
    if above[-1]:
        ends = np.append(ends, len(signal) - 1)

    # Zeitintervalle zusammenbauen
    intervals = [(timestamps[s], timestamps[e]) for s, e in zip(starts, ends)]

    return intervals

def plot_comfort_analysis(df, method_name, signal_names, thresholds, timestamps, discomfort_intervals, output_dir):
    """
    Plotet pro Methode ein Fenster mit 5 Subplots (je Signal):
    - Signalverlauf
    - rot markierte 'wahre' unkomfortable Zeitbereiche (aus Labels)
    - blau markierte 'detektierte' unkomfortable Zeitbereiche (über Threshold)

    Parameter
    ----------
    df : pandas.DataFrame
        Enthält die Signale und die Spalte 'discomfort' (0/1).
    method_name : str
        Name der Methode ('M1', 'M2', etc.).
    signal_names : list of str
        Liste der Signalnamen im df, die geplottet werden sollen.
    thresholds : dict
        Dictionary {signal_name: threshold_value} für diese Methode.
    timestamps : array-like
        Zeitstempel für die Signale.
    discomfort_intervals : list of tuples
        Manuell erzeugte oder aus df berechnete unkomfortable Zeitintervalle
        z. B. [(start, end), ...] aus Ground Truth.
    """

    fig, axes = plt.subplots(len(signal_names), 1, figsize=(12, 10), sharex=True)
    fig.suptitle(f"Komfortanalyse – {method_name}", fontsize=14)

    for i, signal in enumerate(signal_names):
        ax = axes[i]
        y = df[signal]
        thr = thresholds[signal]

        # Intervall auswählen
        if method_name == 'M2':
            y = df[signal][used_interval]
        elif method_name == 'M3':
            y = df[signal][1.0]

        # Threshold-basierte Erkennung
        detected_intervals = find_threshold_intervals(y, timestamps, thr)
        unkomfortabel_detektiert[method_name] = merge_intervals(unkomfortabel_detektiert[method_name], detected_intervals)

        # Signal plotten
        ax.plot(timestamps, y, label=signal, color='black', lw=1)

        # Ground Truth (rot)
        for (start, end) in discomfort_intervals:
            ax.axvspan(start, end, color='red', alpha=0.25, label='Label (unkomfortabel)')

        # Threshold-Erkennung (blau)
        for (start, end) in detected_intervals:
            ax.axvspan(start, end, color='blue', alpha=0.25, label='Threshold erkannt')

        # Threshold-Linie
        ax.axhline(thr, color='gray', linestyle='--', lw=1, label=f'|Threshold| = {thr:.2f}')
        ax.axhline(-thr, color='gray', linestyle='--', lw=1)

        # y-Label bestimmen
        if 'acc_x' in signal or 'acc_y' in signal:
            ylabel = f"{signal} [m/s²]"
        elif 'acc_yaw' in signal:
            ylabel = f"{signal} [rad/s²]"
        elif 'jerk_x' in signal or 'jerk_y' in signal:
            ylabel = f"{signal} [m/s³]"

        ax.set_ylabel(ylabel)
        ax.grid(True)

        # Doppelte Legenden verhindern
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        ax.legend(by_label.values(), by_label.keys(), loc='upper right', fontsize=8)

    axes[-1].set_xlabel("Zeit [s]")
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    plt.show()
    fig.savefig(os.path.join(output_dir, f"Komfortanalyse_mit_{method_name}.png"))

def merge_intervals(list1, list2):
    """
    Vereinigt zwei Listen von Zeitintervallen.
    Überlappende oder aneinandergrenzende Intervalle werden zusammengeführt.

    Parameter:
        list1, list2: Listen von Tupeln (start, end)

    Rückgabe:
        Liste zusammengeführter Intervalle (start, end)
    """
    # Beide Listen zusammenführen
    intervals = list1 + list2

    # Falls leer, direkt zurückgeben
    if not intervals:
        return []

    # Nach Startzeit sortieren
    intervals.sort(key=lambda x: x[0])

    merged = [intervals[0]]

    for current in intervals[1:]:
        prev_start, prev_end = merged[-1]
        curr_start, curr_end = current

        # Prüfen, ob sich die Intervalle überlappen oder berühren
        if curr_start <= prev_end:
            # Zusammenführen
            merged[-1] = (prev_start, max(prev_end, curr_end))
        else:
            # Kein Überlapp, neues Intervall hinzufügen
            merged.append(current)

    return merged

def evaluate_intervals(labeled_intervals, predicted_intervals, t_total, tolerance=2.5):
    """
    Berechnet eine zeitbasierte Konfusionsmatrix aus gelabelten und vorhergesagten unkomfortablen Intervallen.

    Regeln:
      - Wenn ein predicted-Intervall ein gelabeltes Intervall berührt oder überlappt → TP.
      - Falls das predicted-Intervall mehr als ±tolerance über das gelabelte hinausgeht → 
        die überstehenden Bereiche zählen als FP.
      - Zeitpunkte außerhalb aller predicted-Intervalle, aber innerhalb gelabelter → FN.
      - Zeitpunkte außerhalb beider → TN.

    Parameter:
        labeled_intervals: Liste von (start, end) — echte unkomfortable Zeitintervalle
        predicted_intervals: Liste von (start, end) — vorhergesagte unkomfortable Zeitintervalle
        t_total: float — Gesamtdauer der Fahrt (z. B. Sekunden)
        tolerance: float — maximale erlaubte Differenz (z. B. 2.5 s)

    Rückgabe:
        dict mit { 'TP': Dauer, 'FP': Dauer, 'FN': Dauer, 'TN': Dauer }
    """
    TP, FP, FN = 0.0, 0.0, 0.0

    # Hilfsfunktion: prüft, ob Intervalle sich überschneiden
    def overlaps(a, b):
        return not (a[1] <= b[0] or b[1] <= a[0])

    for p_start, p_end in predicted_intervals:
        overlap_found = False
        for l_start, l_end in labeled_intervals:
            if overlaps((p_start, p_end), (l_start, l_end)):
                overlap_found = True
                # TP-Bereich erweitern (± tolerance)
                tp_start = max(p_start, l_start - tolerance)
                tp_end = min(p_end, l_end + tolerance)
                tp_duration = max(0, tp_end - tp_start)
                TP += tp_duration

                # Bereich außerhalb der Toleranz als FP
                if p_start < tp_start:
                    FP += tp_start - p_start
                if p_end > tp_end:
                    FP += p_end - tp_end
                break  # reicht, wenn mindestens ein Label-Intervall überlappt
        if not overlap_found:
            # Kein Label im predicted Intervall → komplett FP
            FP += p_end - p_start

    # Nun FN berechnen: Label-Bereiche, die von keinem Predicted überlappt werden
    for l_start, l_end in labeled_intervals:
        matched = any(overlaps((l_start, l_end), (p_start, p_end)) for p_start, p_end in predicted_intervals)
        if not matched:
            FN += l_end - l_start

    # TN = Gesamtdauer - TP - FP - FN
    TN = max(0, t_total - (TP + FP + FN))

    return {"TP": TP, "FP": FP, "FN": FN, "TN": TN}

def evaluate_metrics(confusion_dict, output_dir, method_name, show_plot=True):
    """
    Berechnet Precision, Recall, F1-Score, Accuracy und zeigt optional eine Konfusionsmatrix.

    Parameter:
        confusion_dict: dict mit { 'TP': ..., 'FP': ..., 'FN': ..., 'TN': ... }
        show_plot: bool – ob eine visuelle Konfusionsmatrix geplottet werden soll

    Rückgabe:
        dict mit allen Metriken
    """
    TP = confusion_dict['TP']
    FP = confusion_dict['FP']
    FN = confusion_dict['FN']
    TN = confusion_dict['TN']

    # --- Berechnung der Metriken ---
    precision = TP / (TP + FP) if (TP + FP) > 0 else 0
    recall = TP / (TP + FN) if (TP + FN) > 0 else 0
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
    accuracy = (TP + TN) / (TP + TN + FP + FN) if (TP + TN + FP + FN) > 0 else 0

    metrics = {
        "Precision": precision,
        "Recall": recall,
        "F1-Score": f1,
        "Accuracy": accuracy
    }

    # --- Plot der Konfusionsmatrix ---
    if show_plot:
        cm = np.array([[TP, FP],
                       [FN, TN]])
        labels = [["TP", "FP"],
                  ["FN", "TN"]]

        fig, ax = plt.subplots(figsize=(5.6, 4))
        im = ax.imshow(cm, cmap="Blues")

        # Werte anzeigen
        for i in range(2):
            for j in range(2):
                ax.text(j, i, f"{labels[i][j]}\n{cm[i, j]:.2f}",
                        ha="center", va="center", color="black", fontsize=11)

        ax.set_xticks([0, 1])
        ax.set_yticks([0, 1])
        ax.set_xticklabels(["Positiv vorhergesagt", "Negativ vorhergesagt"])
        ax.set_yticklabels(["Tatsächlich positiv", "Tatsächlich negativ"])
        ax.set_title(f"Zeitbasierte Konfusionsmatrix - {method_name}")
        plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
        plt.tight_layout()
        plt.show()
        fig.savefig(os.path.join(output_dir, f"Konfusionsmatrix_von_{method_name}.png"))

    return metrics

def compare_methods(metrics_dict, output_dir, show_values=True):
    """
    Vergleicht mehrere Methoden anhand ihrer Metriken (Precision, Recall, F1, Accuracy)
    und plottet ein gruppiertes Balkendiagramm.

    Parameter:
        metrics_dict: dict
            Struktur: {
                'M1': {'Precision': ..., 'Recall': ..., 'F1-Score': ..., 'Accuracy': ...},
                'M2': {...},
                ...
            }
        show_values: bool
            Ob die Werte oberhalb der Balken angezeigt werden sollen.
    """
    methods = list(metrics_dict.keys())
    metrics = list(next(iter(metrics_dict.values())).keys())

    # Werte in eine Matrix konvertieren
    values = np.array([[metrics_dict[m][metric] for metric in metrics] for m in methods])

    x = np.arange(len(metrics))
    width = 0.15  # Breite der Balken

    fig, ax = plt.subplots(figsize=(10, 5))

    # Farben aus colormap generieren
    colors = plt.cm.tab10(np.linspace(0, 1, len(methods)))

    # Balken plotten
    for i, (method, color) in enumerate(zip(methods, colors)):
        ax.bar(x + i * width, values[i], width, label=method, color=color, alpha=0.8)

        # Werte über den Balken anzeigen
        if show_values:
            for j, val in enumerate(values[i]):
                ax.text(x[j] + i * width, val + 0.01, f"{val:.2f}", ha='center', va='bottom', fontsize=9)

    ax.set_xticks(x + width * (len(methods) - 1) / 2)
    ax.set_xticklabels(metrics)
    ax.set_ylabel("Wert")
    ax.set_ylim(0, 1.1)
    ax.set_title("Vergleich der Methoden anhand der Klassifikationsmetriken")
    ax.legend(title="Methode")
    plt.tight_layout()
    plt.show()
    fig.savefig(os.path.join(output_dir, f"Vergleich_Analysemetriken.png"))

# -------------------
# CSV laden. WITCHTIG: In der csv-Datei müssen die Spalten richtig benannt sein: timestamp (in Sekunden!), linear_acceleration_x, linear_acceleration_y, angular_velocity_z. Die Reihenfolge und ob noch andere Spalten sind spielt keinen Rolle.
# -------------------
Tk().withdraw()
filename = askopenfilename(
    title="Bitte CSV-Datei auswählen",
    filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
)

df = pd.read_csv(filename)

# relative Zeit (s)
t0 = df["timestamp"].iloc[0]
df["time_rel"] = (df["timestamp"] - t0)
# Kontrolle ob s oder ms
if df['time_rel'][1] > 1:
    df['time_rel'] /= 1000

# Spalten bereinigen & umbenennen.
columns_to_keep = ["timestamp", "time_rel", "linear_acceleration_x", "linear_acceleration_y", "angular_velocity_z"]
df = df[columns_to_keep]
df = df.rename(columns={
    "linear_acceleration_x": "acc_x",
    "linear_acceleration_y": "acc_y",
    "angular_velocity_z": "yaw"
})

# -------------------
# Abtastrate wählen (EDGAR: 25ms, HyperIMU: 100ms)
# -------------------
# Hauptfenster
root = tk.Tk()
root.title("Abtastrate wählen")

# Label
label = tk.Label(root, text="Bitte Abtastrate (in ms) wählen.\nEDGAR: 25ms\nHyperIMU: 100ms")
label.pack(padx=10, pady=10)

# Dropdown-Menü (Combobox)
options = [25.0, 100.0]
combo = ttk.Combobox(root, values=options, state="readonly")
combo.current(0)  # Standardwert
combo.pack(padx=10, pady=5)

# OK-Button
ok_button = tk.Button(root, text="OK", command=confirm_selection)
ok_button.pack(padx=10, pady=10)

# GUI starten
root.wait_window()

# =====================================================
# LABEL ERSTELLEN
# =====================================================
df['discomfort'] = 0  # Default = komfortabel

for start, end in unangenehme_intervalle:
    df.loc[(df['time_rel'] >= start) & (df['time_rel'] <= end), 'discomfort'] = 1

print(f"{df['discomfort'].sum()} Zeitpunkte als 'unkomfortabel' markiert.")

# ==================================================================================================
# SIGNALBERECHNUNGEN (M1-gefilterte Signale, M2-Ausummierte Signale, M3-RMS Berechnung nach ISO 2631
# ==================================================================================================
# -------------------
# Filter anwenden
# -------------------
cutoff = 2.0   # Grenzfrequenz [Hz]
df["acc_x_filt"] = lowpass_filter(df["acc_x"], cutoff, fs_glob)
df["acc_y_filt"] = lowpass_filter(df["acc_y"], cutoff, fs_glob)
df['yaw_filt'] = lowpass_filter(df["yaw"], cutoff, fs_glob)

# -------------------
# Winkelbeschleunigung berechnen
# -------------------
a = df['yaw_filt'].values
t = df['time_rel'].values
acc_yaw = np.gradient(a, t)
df['acc_yaw'] = acc_yaw

acc_yaw_cut = 1.0   # Hz
df['acc_yaw_filt'] = lowpass_filter(df['acc_yaw'].values, acc_yaw_cut, fs_glob, order=2)

# -------------------
# Jerk berechnen
# -------------------
for axis in ['x', 'y']:
    a = df[f'acc_{axis}_filt'].values
    t = df['time_rel'].values
    jerk = np.gradient(a, t)
    df[f'jerk_{axis}_raw'] = jerk

jerk_cut = 1.0  # Hz
df["jerk_x_filt"] = lowpass_filter(df["jerk_x_raw"].values, jerk_cut, fs_glob, order=2)
df["jerk_y_filt"] = lowpass_filter(df["jerk_y_raw"].values, jerk_cut, fs_glob, order=2)

# ------------------------------
# Summen für Methode 2 berechnen
# ------------------------------
# Integral/Aufsummierung der Beschleunigungs- und Jerkkurven (Absolutwerte)
sum_x = integral_abs(df['time_rel'], df['acc_x_filt'])
sum_y = integral_abs(df['time_rel'], df['acc_y_filt'])
print(f"Integral x-Beschleunigung: {sum_x}\nIntegral y-Beschleunigung: {sum_y}")
# Intervallweise berechnen (1s, 3s, 5s)
t_total = int(df['time_rel'].iloc[-1])
intervals = [1.0, 3.0, 5.0, t_total]
sum_x, sum_x_interval_centers = {}, {}
sum_y, sum_y_interval_centers = {}, {}
sum_acc_yaw, sum_acc_yaw_interval_centers = {}, {}
sum_jerk_x, sum_jerk_x_interval_centers = {}, {}
sum_jerk_y, sum_jerk_y_interval_centers = {}, {}
for i in intervals:
    temp1, temp2 = sliding_metrics(df['acc_x_filt'], df['time_rel'], i)
    sum_x[i] = temp1
    sum_x_interval_centers[i] = temp2
    temp1, temp2 = sliding_metrics(df['acc_y_filt'], df['time_rel'], i)
    sum_y[i] = temp1
    sum_y_interval_centers[i] = temp2
    temp1, temp2 = sliding_metrics(df['acc_yaw_filt'], df['time_rel'], i)
    sum_acc_yaw[i] = temp1
    sum_acc_yaw_interval_centers[i] = temp2
    temp1, temp2 = sliding_metrics(df['jerk_x_filt'], df['time_rel'], i)
    sum_jerk_x[i] = temp1
    sum_jerk_x_interval_centers[i] = temp2
    temp1, temp2 = sliding_metrics(df['jerk_y_filt'], df['time_rel'], i)
    sum_jerk_y[i] = temp1
    sum_jerk_y_interval_centers[i] = temp2    
sum_all = {'acc_x': sum_x, 'acc_y': sum_y, 'acc_yaw': sum_acc_yaw, 'jerk_x': sum_jerk_x, 'jerk_y': sum_jerk_y}

# --------------------------------------------------------------------------
# ISO 2631 Metrik - Berechnungen(RMS der frequenzgewichteten Beschleunigung)
# --------------------------------------------------------------------------
b, a = design_wd_filter(fs_glob)     # Gewichtung ist laut Norm für die x- und y-Beschleunigung diesselbe
df['acc_x_normfilt'] = filtfilt(b, a, df['acc_x'])
df['acc_y_normfilt'] = filtfilt(b, a, df['acc_y'])
b, a = design_we_filter(fs_glob)
df['acc_yaw_normfilt'] = filtfilt(b, a, df['acc_yaw'])
df['acc_yaw_normfilt'] *= 0.2   # k-Faktor nach Norm [m/rad]

# Berechne die rms Werte in unterschiedlich großen gleitenden Intervallen
# Pro Beschleunigungsrichtung gibt es ein dict welches als Schlüssel die jeweiligen Intervalle hat {1.0: [0.674, 0.121, ...], 3.0: [...], ...}
acc_x_rms_dict, x_interval_centers = {}, {}
acc_y_rms_dict, y_interval_centers = {}, {}
acc_yaw_rms_dict, yaw_interval_centers = {}, {}
for i in intervals:
    temp1, temp2 = sliding_metrics(df['acc_x_normfilt'], df['time_rel'], i, weighted=True)
    acc_x_rms_dict[i] = temp1
    x_interval_centers[i] = temp2
    temp1, temp2 = sliding_metrics(df['acc_y_normfilt'], df['time_rel'], i, weighted=True)
    acc_y_rms_dict[i] = temp1
    y_interval_centers[i] = temp2
    temp1, temp2 = sliding_metrics(df['acc_yaw_normfilt'], df['time_rel'], i, weighted=True)
    acc_yaw_rms_dict[i] = temp1
    yaw_interval_centers[i] = temp2
rms_all = {'acc_x': acc_x_rms_dict, 'acc_y': acc_y_rms_dict, 'acc_yaw': acc_yaw_rms_dict}

# Alternative rms Berechnungen laut Norm 2631
# MTVV (max transient vibration value) aus sliding rms (Tau=1s)
MTVV_x = np.max(acc_x_rms_dict[1.0])
MTVV_y = np.max(acc_y_rms_dict[1.0])
MTVV_yaw = np.max(acc_yaw_rms_dict[1.0])

# VDV (∫(aw^4)dt)^(1/4)
integral = np.trapezoid(df['acc_x_normfilt']**4, df['time_rel'])
vdv_x = integral**(1/4)
integral = np.trapezoid(df['acc_y_normfilt']**4, df['time_rel'])
vdv_y = integral**(1/4)
integral = np.trapezoid(df['acc_yaw_normfilt']**4, df['time_rel'])
vdv_yaw = integral**(1/4)

# Berechnen ob laut Norm die Alternativen rms Methoden besser sind
# MTVV Verhältnis berechnen (Wenn Schwellwert überschritten, macht es mehr Sinn intervallweise (1s) auszuwerten)
ratio_mtvv = MTVV_x/acc_x_rms_dict[t_total]
print(f"MTVV-Verhältnis für x-Beschleunigung: {ratio_mtvv}. (Schwellwert: 1,5)")
ratio_mtvv = MTVV_y/acc_y_rms_dict[t_total]
print(f"MTVV-Verhältnis für y-Beschleunigung: {ratio_mtvv}. (Schwellwert: 1,5)")
ratio_mtvv = MTVV_yaw/acc_yaw_rms_dict[t_total]
print(f"MTVV-Verhältnis für yaw-Beschleunigung: {ratio_mtvv}. (Schwellwert: 1,5)")

print(acc_x_rms_dict[t_total])
ratio_vdv = vdv_x/(acc_x_rms_dict[t_total][0]*(t_total**(1/4)))
print(f"VDV-Verhältnis für x-Beschleunigung: {ratio_vdv}. (Schwellwert: 1,75)")
ratio_vdv = vdv_y/(acc_y_rms_dict[t_total][0]*(t_total**(1/4)))
print(f"VDV-Verhältnis für y-Beschleunigung: {ratio_vdv}. (Schwellwert: 1,75)")
ratio_vdv = vdv_yaw/(acc_yaw_rms_dict[t_total][0]*(t_total**(1/4)))
print(f"VDV-Verhältnis für yaw-Beschleunigung: {ratio_vdv}. (Schwellwert: 1,75)")

# ======================
# SIGNALVISUALISIERUNGEN
# ======================
# Roh vs. gefilterte Signale
fig, axes = plt.subplots(3, 1, figsize=(12, 12), sharex=True)

# 1) Beschleunigung
axes[0].plot(df['time_rel'], df['acc_x'], label='a_x (roh)', alpha=0.5)
axes[0].plot(df['time_rel'], df['acc_x_filt'], label='a_x (gefiltert)', linewidth=2)
axes[0].plot(df['time_rel'], df['acc_y'], label='a_y (roh)', alpha=0.5)
axes[0].plot(df['time_rel'], df['acc_y_filt'], label='a_y (gefiltert)', linewidth=2)
axes[0].set_ylabel('a [m/s²]')
axes[0].set_title('Lineare Beschleunigungen')
axes[0].grid()
axes[0].legend()

# 2) Yaw und Yaw-Beschleunigung
axes[1].plot(df['time_rel'], df['yaw'], label='ω')
axes[1].plot(df['time_rel'], df['acc_yaw'], label='α (roh)')
axes[1].plot(df['time_rel'], df['acc_yaw_filt'], label='α (gefiltert)')
axes[1].set_ylabel('ω/α [rad/s]/[rad/s²]')
axes[1].set_title('Gierrate und -beschleunigung')
axes[1].grid()
axes[1].legend()

# 3) Jerk
axes[2].plot(df['time_rel'], df['jerk_x_filt'], label='j_x (gefiltert)')
axes[2].plot(df['time_rel'], df['jerk_y_filt'], label='j_y (gefiltert)')
axes[2].set_xlabel('Zeit [s]')
axes[2].set_ylabel('j [m/s³]')
axes[2].set_title('Jerk')
axes[2].grid()
axes[2].legend()

plt.tight_layout()
plt.show()

# gefilterte und aufsummierte Signale (M2)
num_plots = len(sum_all)
fig, axes = plt.subplots(num_plots, 1, figsize=(12, 12), sharex=True)

for i, (key, sig) in enumerate(sum_all.items()):
    axes[i].plot(df['time_rel'], df[f'{key}_filt'], label=f'{key} (gefiltert)', linewidth=2)
    for j in intervals[:-1]:
        axes[i].step(sum_x_interval_centers[j], sig[j], where="mid", label=f"{j:.0f}s Intervall", linewidth=2)
    axes[i].plot()
    match key:
        case 'acc_yaw':
            axes[i].set_ylabel('α [rad/s²]')
        case 'jerk_x' | 'jerk_y':
            axes[i].set_ylabel('j [m/s³]')
        case _:
            axes[i].set_ylabel('a [m/s²]')
    axes[i].set_title(f'Summation des {key} Signals nach Intervallgrößen')
    axes[i].grid()
    axes[i].legend()

plt.tight_layout()
plt.show()

# Gefilterte Beschleunigungsgrafiken (M3)
fig, axes = plt.subplots(3, 1, figsize=(12, 12), sharex=True)

axes[0].plot(df['time_rel'], df['acc_x'], label='Roh', alpha=0.7)
axes[0].plot(df['time_rel'], df['acc_x_normfilt'], label='Gefiltert', alpha = 0.7)
for i in intervals[:-1]:
    axes[0].step(x_interval_centers[i], acc_x_rms_dict[i], where="mid", label=f"RMS {i:.0f}s Intervall", linewidth=2)
axes[0].plot()
axes[0].set_ylabel('a_x [m/s²]')
axes[0].set_title('Norm-gefilterte Längsbeschleunigung')
axes[0].grid()
axes[0].legend()

axes[1].plot(df['time_rel'], df['acc_y'], label='Roh', alpha=0.7)
axes[1].plot(df['time_rel'], df['acc_y_normfilt'], label='Gefiltert', alpha = 0.7)
for i in intervals[:-1]:
    axes[1].step(y_interval_centers[i], acc_y_rms_dict[i], where="mid", label=f"RMS {i:.0f}s Intervall", linewidth=2)
axes[1].set_ylabel('a_y [m/s²]')
axes[1].set_title('Norm-gefilterte Querbeschleunigung')
axes[1].grid()
axes[1].legend()

axes[2].plot(df['time_rel'], df['acc_yaw'], label='Roh', alpha=0.7)
axes[2].plot(df['time_rel'], df['acc_yaw_normfilt'], label='Gefiltert', alpha = 0.7)
for i in intervals[:-1]:
    axes[2].step(yaw_interval_centers[i], acc_yaw_rms_dict[i], where="mid", label=f"RMS {i:.0f}s Intervall", linewidth=2)
axes[2].set_ylabel('α [rad/s²]')
axes[2].set_xlabel('Zeit [s]')
axes[2].set_title('Norm-gefilterte Gierbeschleunigung')
axes[2].grid()
axes[2].legend()

plt.tight_layout()
plt.show()

# =============================================
# PLOTS FÜR DATENBESCHAFFENHEIT/-CHARAKTERISTIK
# =============================================
# ---------------------------
# Speichereinstellungen
# ---------------------------
plt.rcParams.update({'font.size': 10})
output_dir = "Ergebnisse"
os.makedirs(output_dir, exist_ok=True)

# ---------------------------
# Speicherordner festlegen
# ---------------------------
# Fenster verstecken (wir brauchen nur den Dialog)
root = tk.Tk()
root.withdraw()

# Dialog zum Auswählen eines Ordners öffnen
output_dir = filedialog.askdirectory(title="Bitte Speicherort auswählen")

# Abbrechen-Check
if not output_dir:
    print("Kein Speicherort ausgewählt. Programm wird beendet.")
    exit()

print(f"Gewählter Speicherort: {output_dir}")

# ===========================
# 1. Signalverläufe
# ===========================
fig, axs = plt.subplots(5, 1, figsize=(12, 12), sharex=True)
signals = [
    ('acc_x', 'acc_x_filt', 'acc_x_normfilt', 'a_x [m/s²]', 'a_x'),
    ('acc_y', 'acc_y_filt', 'acc_y_normfilt', 'a_y [m/s²]', 'a_y'),
    ('acc_yaw', 'acc_yaw_filt', 'acc_yaw_normfilt', 'α [rad/s²]', 'α'),
    ('jerk_x_raw', 'jerk_x_filt', 'undefined', 'j_x [m/s³]', 'j_x'),
    ('jerk_y_raw', 'jerk_y_filt', 'undefined', 'j_y [m/s³]', 'j_y')
]

for i, (raw, filt, norm, label, sig_name) in enumerate(signals):
    axs[i].plot(df['time_rel'], df[raw], alpha=0.5, label=f'(Roh)')
    axs[i].plot(df['time_rel'], df[filt], label=f'(Butterworth-Filter)')
    if norm != 'undefined':
        axs[i].plot(df['time_rel'], df[norm], label=f'(Norm-Filter)')
    axs[i].set_ylabel(label)
    axs[i].legend()

axs[-1].set_xlabel("Zeit [s]")
plt.suptitle("Signalverläufe über Zeit")
plt.tight_layout(rect=[0, 0, 1, 0.97])
plt.show()
fig.savefig(os.path.join(output_dir, "01_Signalverläufe.png"))

# ===========================
# 2. Histogramme
# ===========================
# Gefilterte Beschleunigungen/Jerk
fig, axs = plt.subplots(1, 5, figsize=(14, 4))
axes = ['acc_x_filt', 'acc_y_filt', 'acc_yaw_filt', 'jerk_x_filt', 'jerk_y_filt']
titles = ['Längsbeschleunigung', 'Querbeschleunigung', 'Gierbeschleunigung', 'Längsjerk', 'Querjerk']

for ax, col, title in zip(axs, axes, titles):
    # Histogramm berechnen
    counts, bins, patches = ax.hist(df[col], bins=np.linspace(df[col].min(), df[col].max(), 20), color='skyblue', edgecolor='black')
    # Relative Häufigkeiten berechnen (Alle Balken zusammen = 1)
    counts_rel = counts / np.sum(counts)
    # Original-Histogramm löschen
    ax.cla()
    # Neues Histogramm mit relativen Häufigkeiten zeichnen
    ax.bar(bins[:-1], counts_rel, width=np.diff(bins), align='edge', color='skyblue', edgecolor='black')
    ax.set_title(title)
    if col == 'acc_yaw_filt':
        ax.set_xlabel("Beschleunigungsbereich [rad/s²]")
    elif col == 'jerk_x_filt' or col == 'jerk_y_filt':
        ax.set_xlabel("Jerkbereich [m/s³]")
    else:
        ax.set_xlabel("Beschleunigungsbereich [m/s²]")
    ax.set_ylabel("Auftretenshäufigkeit")

plt.suptitle("Histogramme der gefilterten Beschleunigungen")
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.show()
fig.savefig(os.path.join(output_dir, "02_Histogramme_Filt_Acc.png"))

# Normgefilterte Beschleunigungen/Jerk
fig, axs = plt.subplots(1, 3, figsize=(14, 4))
axes = ['acc_x_normfilt', 'acc_y_normfilt', 'acc_yaw_normfilt']
titles = ['Längsbeschleunigung', 'Querbeschleunigung', 'Gierbeschleunigung']

for ax, col, title in zip(axs, axes, titles):
    counts, bins, patches = ax.hist(df[col], bins=np.linspace(df[col].min(), df[col].max(), 20), color='skyblue', edgecolor='black')
    # Relative Häufigkeiten berechnen (Alle Balken zusammen = 1)
    counts_rel = counts / np.sum(counts)
    # Original-Histogramm löschen
    ax.cla()
    # Neues Histogramm mit relativen Häufigkeiten zeichnen
    ax.bar(bins[:-1], counts_rel, width=np.diff(bins), align='edge', color='skyblue', edgecolor='black')
    
    ax.set_title(title)
    if col == 'acc_yaw_normfilt':
        ax.set_xlabel("Beschleunigungsbereich [rad/s²]")
    elif col == 'jerk_x_normfilt' or col == 'jerk_y_filt':
        ax.set_xlabel("Jerkbereich [m/s³]")
    else:
        ax.set_xlabel("Beschleunigungsbereich [m/s²]")
    ax.set_ylabel("Auftretenshäufigkeit")
    ax.set_ylabel("Auftretenshäufigkeit")

plt.suptitle("Histogramme der nach ISO 2631 gefilterten Beschleunigungen")
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.show()
fig.savefig(os.path.join(output_dir, "02_Histogramme_Normfilt_Acc.png"))

titles = [1.0, 3.0, 5.0]
# Summation (Methode 2)
for direction in sum_all.keys():
    fig, axs = plt.subplots(1, 3, figsize=(14, 4))
    for ax, title in zip(axs, titles):
        data = sum_all[direction][title]
        counts, bins, patches = ax.hist(data, bins=np.linspace(np.min(data), np.max(data), 20), color='skyblue', edgecolor='black')
        # Relative Häufigkeiten berechnen (Alle Balken zusammen = 1)
        counts_rel = counts / np.sum(counts)
        # Original-Histogramm löschen
        ax.cla()
        # Neues Histogramm mit relativen Häufigkeiten zeichnen
        ax.bar(bins[:-1], counts_rel, width=np.diff(bins), align='edge', color='skyblue', edgecolor='black')

        ax.set_title(f'{title:.0f}s Intervall')
        if direction == 'acc_yaw':
            ax.set_xlabel("Beschleunigungsbereich [rad/s²]")
        elif direction == 'jerk_x' or direction == 'jerk_y':
            ax.set_xlabel("Jerkbereich [m/s³]")
        else:
            ax.set_xlabel("Beschleunigungsbereich [m/s²]")
        ax.set_ylabel("Auftretenshäufigkeit")

    plt.suptitle(f"Histogramme der Summationsmethode (Methode 2): {direction}-Richtung")
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show()
    fig.savefig(os.path.join(output_dir, f"02_Histogramme_sum_{direction}.png"))

# RMS Werte
for direction in rms_all.keys():
    fig, axs = plt.subplots(1, 3, figsize=(14, 4))
    for ax, title in zip(axs, titles):
        data = rms_all[direction][title]
        counts, bins, patches = ax.hist(data, bins=np.linspace(np.min(data), np.max(data), 20), color='skyblue', edgecolor='black')
        # Relative Häufigkeiten berechnen (Alle Balken zusammen = 1)
        counts_rel = counts / np.sum(counts)
        # Original-Histogramm löschen
        ax.cla()
        # Neues Histogramm mit relativen Häufigkeiten zeichnen
        ax.bar(bins[:-1], counts_rel, width=np.diff(bins), align='edge', color='skyblue', edgecolor='black')
        ax.set_title(f'{title:.0f}s Intervall')
        if direction == 'acc_yaw':
            ax.set_xlabel("Beschleunigungsbereich [rad/s²]")
        else:
            ax.set_xlabel("Beschleunigungsbereich [m/s²]")
        ax.set_ylabel("Auftretenshäufigkeit")

    plt.suptitle(f"Histogramme der RMS-Methode (Methode 3): {direction}-Richtung")
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.show()
    fig.savefig(os.path.join(output_dir, f"02_Histogramme_rms_{direction}.png"))

# ===========================
# 3. Summenplots
# ===========================
fig, axs = plt.subplots(5, 1, figsize=(12, 8), sharex=True)
for i, (sum_dict, center_dict, title) in enumerate([
    (sum_x, sum_x_interval_centers, 'a_x [m/s²]'),
    (sum_y, sum_y_interval_centers, 'a_y [m/s²]'),
    (sum_acc_yaw, sum_acc_yaw_interval_centers, 'α [rad/s²]'),
    (sum_jerk_x, sum_jerk_x_interval_centers, 'j_x [m/s³]'),
    (sum_jerk_y, sum_jerk_y_interval_centers, 'j_y [m/s³]')
]):
    for k in intervals[:-1]:
        axs[i].plot(center_dict[k], sum_dict[k], label=f"{k:.0f}s Intervall")
    axs[i].set_ylabel(title)
    axs[i].legend()
axs[-1].set_xlabel("Zeit [s]")
plt.suptitle("Methode 2: Aufsummierte Beschleunigungen/Jerks")
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.show()
fig.savefig(os.path.join(output_dir, "03_Summenplots.png"))

# ===========================
# 4. RMS-Plots
# ===========================
fig, axs = plt.subplots(3, 1, figsize=(12, 8), sharex=True)
for i, (rms_dict, centers, title) in enumerate([
    (acc_x_rms_dict, x_interval_centers, 'x-RMS [m/s²]'),
    (acc_y_rms_dict, y_interval_centers, 'y-RMS [m/s²]'),
    (acc_yaw_rms_dict, yaw_interval_centers, 'α-RMS [rad/s²]')
]):
    for k in intervals[:-1]:
        axs[i].plot(centers[k], rms_dict[k], label=f"{k:.0f}s Intervall")
    axs[i].set_ylabel(title)
    axs[i].legend()
axs[-1].set_xlabel("Zeit [s]")
plt.suptitle("Methode 3: RMS-Beschleunigungen")
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.show()
fig.savefig(os.path.join(output_dir, "04_RMS_Plots.png"))

# ===========================
# 5. Box-Plots
# ===========================
labels = ['Roh', 'Butterworth', 'Norm', 'Sum_1s', 'Sum_3s', 'Sum_5s', 'RMS_1s', 'RMS_3s', 'RMS_5s']
labels_jerk = ['Roh', 'Butterworth', 'Sum_1s', 'Sum_3s', 'Sum_5s']

# Längsbeschleunigung
data_box_acc_x = [
    df['acc_x'], df['acc_x_filt'], df['acc_x_normfilt'],
    sum_x[1.0], sum_x[3.0], sum_x[5.0], acc_x_rms_dict[1.0], acc_x_rms_dict[3.0], acc_x_rms_dict[5.0]
]

# Querbeschleunigung
data_box_acc_y = [
    df['acc_y'], df['acc_y_filt'], df['acc_y_normfilt'],
    sum_y[1.0], sum_y[3.0], sum_y[5.0], acc_y_rms_dict[1.0], acc_y_rms_dict[3.0], acc_y_rms_dict[5.0]
]

# Gierbeschleunigung
data_box_acc_yaw = [
    df['acc_yaw'], df['acc_yaw_filt'], df['acc_yaw_normfilt'],
    sum_acc_yaw[1.0], sum_acc_yaw[3.0], sum_acc_yaw[5.0], acc_yaw_rms_dict[1.0], acc_yaw_rms_dict[3.0], acc_yaw_rms_dict[5.0]
]

# Längsjerk
data_box_jerk_x = [
    df['jerk_x_raw'], df['jerk_x_filt'], sum_jerk_x[1.0], sum_jerk_x[3.0], sum_jerk_x[5.0]
]

# Querjerk
data_box_jerk_y = [
    df['jerk_x_raw'], df['jerk_x_filt'], sum_jerk_x[1.0], sum_jerk_x[3.0], sum_jerk_x[5.0]
]

fig, ax = plt.subplots(figsize=(10, 5))
ax.boxplot(data_box_acc_x, labels=labels)
ax.set_title("Vergleich aller x-Beschleunigungsdaten")
ax.set_ylabel("Beschleunigung [m/s²]")
plt.show()
fig.savefig(os.path.join(output_dir, "05_Boxplots_acc_x.png"))

fig, ax = plt.subplots(figsize=(10, 5))
ax.boxplot(data_box_acc_y, labels=labels)
ax.set_title("Vergleich aller y-Beschleunigungsdaten")
ax.set_ylabel("Beschleunigung [m/s²]")
plt.show()
fig.savefig(os.path.join(output_dir, "05_Boxplots_acc_y.png"))

fig, ax = plt.subplots(figsize=(10, 5))
ax.boxplot(data_box_acc_yaw, labels=labels)
ax.set_title("Vergleich aller Gier-Beschleunigungsdaten")
ax.set_ylabel("Beschleunigung [rad/s²]")
plt.show()
fig.savefig(os.path.join(output_dir, "05_Boxplots_acc_yaw.png"))

fig, ax = plt.subplots(figsize=(10, 5))
ax.boxplot(data_box_jerk_x, labels=labels_jerk)
ax.set_title("Vergleich aller Jerkdaten in x-Richtung")
ax.set_ylabel("Jerk [m/s³]")
plt.show()
fig.savefig(os.path.join(output_dir, "05_Boxplots_jerk_x.png"))

fig, ax = plt.subplots(figsize=(10, 5))
ax.boxplot(data_box_jerk_y, labels=labels_jerk)
ax.set_title("Vergleich aller Jerkdaten in y-Richtung")
ax.set_ylabel("Jerk [m/s³]")
plt.show()
fig.savefig(os.path.join(output_dir, "05_Boxplots_jerk_y.png"))

# ===========================
# 6. Heatmap
# ===========================
sum_x_means, sum_y_means, sum_acc_yaw_means, sum_jerk_x_means, sum_jerk_y_means = {}, {}, {}, {}, {}
rms_x_means, rms_y_means, rms_yaw_means = {}, {}, {}
for i in intervals[:-1]:
    sum_x_means[i] = np.mean(sum_x[i])
    sum_y_means[i] = np.mean(sum_y[i])
    sum_acc_yaw_means[i] = np.mean(sum_acc_yaw[i])
    sum_jerk_x_means[i] = np.mean(sum_jerk_x[i])
    sum_jerk_y_means[i] = np.mean(sum_jerk_x[i])
    rms_x_means[i] = np.mean(acc_x_rms_dict[i])
    rms_y_means[i] = np.mean(acc_y_rms_dict[i])
    rms_yaw_means[i] = np.mean(acc_yaw_rms_dict[i])
heat_data = pd.DataFrame({
    'Roh': [df['acc_x'].std(), df['acc_y'].std(), df['acc_yaw'].std(), df['jerk_x_raw'].std(), df['jerk_y_raw'].std()],
    'Filt': [df['acc_x_filt'].std(), df['acc_y_filt'].std(), df['acc_yaw_filt'].std(), df['jerk_x_filt'].std(), df['jerk_y_filt'].std()],
    'Norm': [df['acc_x_normfilt'].std(), df['acc_y_normfilt'].std(), df['acc_yaw_normfilt'].std(), None, None],
    'Sum_1s': [sum_x_means[intervals[0]], sum_y_means[intervals[0]], sum_acc_yaw_means[intervals[0]], sum_jerk_x_means[intervals[0]], sum_jerk_y_means[intervals[0]]],
    'Sum_3s': [sum_x_means[intervals[1]], sum_y_means[intervals[1]], sum_acc_yaw_means[intervals[1]], sum_jerk_x_means[intervals[1]], sum_jerk_y_means[intervals[1]]],
    'Sum_5s': [sum_x_means[intervals[2]], sum_y_means[intervals[2]], sum_acc_yaw_means[intervals[2]], sum_jerk_x_means[intervals[2]], sum_jerk_y_means[intervals[2]]],
    'RMS_1s': [rms_x_means[intervals[0]], rms_y_means[intervals[0]], rms_yaw_means[intervals[0]], None, None],
    'RMS_3s': [rms_x_means[intervals[1]], rms_y_means[intervals[1]], rms_yaw_means[intervals[1]], None, None],
    'RMS_5s': [rms_x_means[intervals[2]], rms_y_means[intervals[2]], rms_yaw_means[intervals[2]], None, None]
}, index=['a_x [m/s²]', 'a_y [m/s²]', 'α [rad/s²]', 'j_x [m/s³]', 'j_y [m/s³]'])

fig, ax = plt.subplots(figsize=(8, 5))
sns.heatmap(heat_data, annot=True, cmap="coolwarm", fmt=".2f")
ax.set_title("Heatmap: Vergleich der Durchschnittswerte")
plt.show()
fig.savefig(os.path.join(output_dir, "06_Heatmaps_all.png"))

# ===========================
# 7. Scatterplots
# ===========================
fig, axs = plt.subplots(2, 2, figsize=(10, 8))
pairs = [
    ('acc_x_filt', 'acc_y_filt', 'x vs y (Butterworth)'),
    ('acc_x_filt', 'acc_yaw_filt', 'x vs yaw (Butterworth)'),
    (list(acc_x_rms_dict.values())[0], list(acc_y_rms_dict.values())[0], 'RMS x vs y'),
    (list(sum_x.values())[0], list(sum_y.values())[0], 'Sum x vs y')
]

for ax, (xdata, ydata, title) in zip(axs.ravel(), pairs):
    if isinstance(xdata, str):
        ax.scatter(df[xdata], df[ydata], alpha=0.5)
    else:
        ax.scatter(xdata, ydata, alpha=0.5)
    ax.set_title(title)
plt.suptitle("Scatterplots – Zusammenhang zwischen Achsen & Methoden")
plt.tight_layout(rect=[0, 0, 1, 0.95])
plt.show()
fig.savefig(os.path.join(output_dir, "07_Scatterplots.png"))

# ================================
# KOMFORTANALYSE
# ================================
# Signaldarstellung Vergleich gelabelte und berechnete unkomfortable Bereiche
# M1
signal_names = ['acc_x_filt', 'acc_y_filt', 'acc_yaw_filt', 'jerk_x_filt', 'jerk_y_filt']
thresholds = {}
for sig in signal_names:
    for sig_n, sig_thresh in thresholds_dict.items():
        if sig_n in sig:
            thresholds[sig] = sig_thresh['m1'] 

plot_comfort_analysis(df, 'M1', signal_names, thresholds, df['time_rel'], unangenehme_intervalle, output_dir)

# M2
if used_interval == 1.0:
    idx = 0
elif used_interval == 3.0:
    idx = 1
elif used_interval == 5.0:
    idx = 2
signal_names = ['acc_x', 'acc_y', 'acc_yaw', 'jerk_x', 'jerk_y']
for sig in signal_names:
    for sig_n, sig_thresh in thresholds_dict.items():
        if sig_n in sig:
            thresholds[sig] = sig_thresh['m2'][idx]

plot_comfort_analysis(sum_all, 'M2', signal_names, thresholds, sum_x_interval_centers[used_interval], unangenehme_intervalle, output_dir)

# M3
signal_names = ['acc_x', 'acc_y', 'acc_yaw']
for sig in signal_names:
    for sig_n, sig_thresh in thresholds_dict.items():
        if sig_n in sig:
            thresholds[sig] = sig_thresh['m3'][3]
plot_comfort_analysis(rms_all, 'M3', signal_names, thresholds, sum_x_interval_centers[1.0], unangenehme_intervalle, output_dir)

# Konfusionsmatrix und klassische Kennzahlen (Precision, Recall, F1)
confusion_dict = {}
metrics_dict = {}
t_total = df['time_rel'].iloc[-1]
for meth in ['M1', 'M2', 'M3']:
    confusion_dict[meth] = evaluate_intervals(unangenehme_intervalle, unkomfortabel_detektiert[meth], t_total)
    metrics_dict[meth] = evaluate_metrics(confusion_dict[meth], output_dir, meth)

compare_methods(metrics_dict, output_dir)


# # ================================
# # EXCEL GENERIEREN UND ABSPEICHERN
# # ================================

# max_x = np.asarray(list(get_max_per_interval(df['time_rel'], df['acc_x_filt']).values()))
# max_y = np.asarray(list(get_max_per_interval(df['time_rel'], df['acc_y_filt']).values()))
# max_yaw = np.asarray(list(get_max_per_interval(df['time_rel'], df['yaw_filt']).values()))
# max_jerk_x = np.asarray(list(get_max_per_interval(df['time_rel'], df['jerk_x_filt']).values()))
# max_jerk_y = np.asarray(list(get_max_per_interval(df['time_rel'], df['jerk_y_filt']).values()))
# t_total_int = len(max_x)

# # Signale und ihre Methode1 Werte
# signals_m1 = {
#     "acc_x": max_x,
#     "acc_y": max_y,
#     "acc_yaw": max_yaw
# }

# # Methode 2 Werte (Dicts)
# signals_m2 = {
#     "acc_x": sum_x,
#     "acc_y": sum_y,
#     "acc_yaw": sum_acc_yaw
# }

# # Methode 3 Werte (Dicts)
# signals_m3 = {
#     "acc_x": acc_x_rms_dict,
#     "acc_y": acc_y_rms_dict,
#     "acc_yaw": acc_yaw_rms_dict
# }

# # === Farben definieren ===
# red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # hellrot
# white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # weiß

# # Neues Workbook erstellen
# wb = openpyxl.Workbook()
# first = True

# for signal_name in ["acc_x", "acc_y", "acc_yaw"]:
#     if first:
#         ws = wb.active
#         ws.title = signal_name
#         first = False
#     else:
#         ws = wb.create_sheet(title=signal_name)
    
#     # ---- Spalte 1: Methoden ----
#     ws.cell(row=1, column=1, value="Zeitleiste in 1s Intervallen")
#     ws.cell(row=2, column=1, value="Methode 1 (Maxwerte pro Intervall)")
#     ws.cell(row=3, column=1, value="Methode 2 (1s Intervall)")
#     ws.cell(row=4, column=1, value="Methode 2 (1s Intervall)")
#     ws.cell(row=5, column=1, value="Methode 2 (3s Intervall)")
#     ws.cell(row=6, column=1, value="Methode 2 (3s Intervall)")
#     ws.cell(row=7, column=1, value="Methode 2 (5s Intervall)")
#     ws.cell(row=8, column=1, value="Methode 2 (5s Intervall)")
#     ws.cell(row=9, column=1, value="Methode 2 (t_total)", Alignment='left')
#     ws.cell(row=10, column=1, value="Methode 3 (1s Intervall)")
#     ws.cell(row=11, column=1, value="Methode 3 (1s Intervall)")
#     ws.cell(row=12, column=1, value="Methode 3 (2s Intervall)")
#     ws.cell(row=13, column=1, value="Methode 3 (2s Intervall)")
#     ws.cell(row=14, column=1, value="Methode 3 (5s Intervall)")
#     ws.cell(row=15, column=1, value="Methode 3 (5s Intervall)")
#     ws.cell(row=16, column=1, value="Methode 3 (t_total)")
#     ws.cell(row=17, column=1, value="Jerk (Maxwerte pro Intervall)")


#     # ---- Zeile 1: 1s Intervalle ----
#     for i in range(1, t_total_int*2, 2):
#         ws.merge_cells(start_row=1, start_column=i+1, end_row=1, end_column=min(i+2, t_total_int*2))
#         ws.cell(row=1, column=i+1, value=f"{int(i/2)}-{int(i/2)+1}s")
#         ws.column_dimensions[get_column_letter(i+1)].width = 10
#         ws.column_dimensions[get_column_letter(i+2)].width = 10
#         ws.cell(row=1, column=i+1).alignment = Alignment(horizontal='center')
#     ws.row_dimensions[1].height = 20
    
#     # ---- Zeile 2: Methode 1 ----
#     m1_values = signals_m1[signal_name]
#     for i in range(1, t_total_int*2, 2):
#         ws.merge_cells(start_row=2, start_column=i+1, end_row=2, end_column=min(i+2, t_total_int*2))
#         ws.cell(row=2, column=i+1, value=round(m1_values[int(i/2)], 3))
#         cell = ws.cell(row=2, column=i+1)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m1']:
#             cell.fill = red_fill
    
#     # ---- Methode 2 ----
#     m2_dict = signals_m2[signal_name]
#     # 1s Intervalle
#     col = 2
#     for i in range(1, len(m2_dict[1.0])-1, 2):
#         ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=min(col+1, t_total_int*2))
#         ws.cell(row=3, column=col, value=round(m2_dict[1.0][i], 3))
#         cell = ws.cell(row=3, column=col)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m2'][0]:
#             cell.fill = red_fill
#         ws.merge_cells(start_row=4, start_column=col+1, end_row=4, end_column=min(col+2, t_total_int*2))
#         ws.cell(row=4, column=col+1, value=round(m2_dict[1.0][i+1], 3))
#         cell = ws.cell(row=4, column=col+1)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m2'][0]:
#             cell.fill = red_fill
#         col += 2
#     # 3s Intervalle
#     col = 2
#     for i in range(1, len(m2_dict[3.0])-1, 2):
#         ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=min(col+5, t_total_int*2))
#         ws.cell(row=5, column=col, value=round(m2_dict[3.0][i], 3))
#         cell = ws.cell(row=5, column=col)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m2'][1]:
#             cell.fill = red_fill
#         ws.merge_cells(start_row=6, start_column=col+3, end_row=6, end_column=min(col+8, t_total_int*2))
#         ws.cell(row=6, column=col+3, value=round(m2_dict[3.0][i+1], 3))
#         cell = ws.cell(row=6, column=col+3)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m2'][1]:
#             cell.fill = red_fill
#         col += 6
#     # 5s Intervalle
#     col = 2
#     for i in range(1, len(m2_dict[5.0])-1, 2):
#         ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=min(col+9, t_total_int*2))
#         ws.cell(row=7, column=col, value=round(m2_dict[5.0][i], 3))
#         cell = ws.cell(row=7, column=col)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m2'][2]:
#             cell.fill = red_fill
#         ws.merge_cells(start_row=8, start_column=col+5, end_row=8, end_column=min(col+14, t_total_int*2))
#         ws.cell(row=8, column=col+5, value=round(m2_dict[5.0][i+1], 3))
#         cell = ws.cell(row=8, column=col+5)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m2'][2]:
#             cell.fill = red_fill
#         col += 10
#     # t_total
#     ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=t_total_int*2)
#     ws.cell(row=9, column=2, value=round(m2_dict[t_total][0], 3))
#     cell = ws.cell(row=9, column=1)
#     cell.alignment = Alignment(horizontal='center')
    
#     # ---- Methode 3 ----
#     m3_dict = signals_m3[signal_name]
#     # 1s Intervalle
#     col = 2
#     for i in range(1, len(m3_dict[1.0])-1, 2):
#         ws.merge_cells(start_row=10, start_column=col, end_row=10, end_column=min(col+1, t_total_int*2))
#         ws.cell(row=10, column=col, value=round(m3_dict[1.0][i], 3))
#         cell = ws.cell(row=10, column=col)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#         ws.merge_cells(start_row=11, start_column=col+1, end_row=11, end_column=min(col+2, t_total_int*2))
#         ws.cell(row=11, column=col+1, value=round(m3_dict[1.0][i+1], 3))
#         cell = ws.cell(row=11, column=col+1)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#         col += 2
#     # 3s Intervalle
#     col = 2
#     for i in range(1, len(m3_dict[3.0])-1, 2):
#         ws.merge_cells(start_row=12, start_column=col, end_row=12, end_column=min(col+5, t_total_int*2))
#         ws.cell(row=12, column=col, value=round(m3_dict[3.0][i], 3))
#         cell = ws.cell(row=12, column=col)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#         ws.merge_cells(start_row=13, start_column=col+3, end_row=13, end_column=min(col+8, t_total_int*2))
#         ws.cell(row=13, column=col+3, value=round(m3_dict[3.0][i+1], 3))
#         cell = ws.cell(row=13, column=col+3)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#         col += 6
#     # 5s Intervalle
#     col = 2
#     for i in range(1, len(m3_dict[5.0])-1, 2):
#         ws.merge_cells(start_row=14, start_column=col, end_row=14, end_column=min(col+9, t_total_int*2))
#         ws.cell(row=14, column=col, value=round(m3_dict[5.0][i], 3))
#         cell = ws.cell(row=14, column=col)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#         ws.merge_cells(start_row=15, start_column=col+5, end_row=15, end_column=min(col+14, t_total_int*2))
#         ws.cell(row=15, column=col+5, value=round(m3_dict[5.0][i+1], 3))
#         cell = ws.cell(row=15, column=col+5)
#         cell.alignment = Alignment(horizontal='center')
#         if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#         col += 10
#     # t_total
#     ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=t_total_int*2)
#     ws.cell(row=16, column=2, value=round(m3_dict[t_total][0], 3))
#     cell = ws.cell(row=16, column=2)
#     cell.alignment = Alignment(horizontal='center')
#     if cell.value > thresholds_dict[signal_name]['m3'][3]:
#             cell.fill = red_fill
#     # Jerk
#     if signal_name == 'acc_x' or signal_name == 'acc_y':
#         if signal_name == 'acc_x':
#             sig = 'jerk_y'
#             jerk_max_interval = max_jerk_x
#         else:
#             sig = 'jerk_y'
#             jerk_max_interval = max_jerk_y
#         for i in range(1, t_total_int*2, 2):
#             ws.merge_cells(start_row=17, start_column=i+1, end_row=17, end_column=min(i+2, t_total_int*2))
#             ws.cell(row=17, column=i+1, value=round(jerk_max_interval[int(i/2)], 3))
#             cell = ws.cell(row=17, column=i+1)
#             cell.alignment = Alignment(horizontal='center')
#             if cell.value > thresholds_dict[sig]:
#                 cell.fill = red_fill

# # Excel speichern
# #wb.save(f'{output_dir}\Signalvergleich.xlsx')    # C:\Users\kompa\Documents\University\TUM\Bachelor Thesis\Excel Analysis\IMU_Methoden_Vergleich_Signale.xlsx
# #print("Excel-Datei 'IMU_Methoden_Vergleich_Signale.xlsx' erstellt!")

